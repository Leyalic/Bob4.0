"""
FileProcessor: all file I/O, year detection, routing, copy/move logic.
No tkinter imports — this module is purely business logic.
"""
from __future__ import annotations

import calendar
import csv
import datetime
import io
import logging
import os
import re
import shutil
import time
import warnings
from pathlib import Path
from typing import Optional

import openpyxl

from Files import config
from Files.routes import ROUTES

# ── Logging ────────────────────────────────────────────────────────────────
config.LOG_DIR.mkdir(exist_ok=True)
LOG_FILE = config.LOG_DIR / f"bob_{time.strftime('%Y%m%d_%H%M%S')}.log"
logging.basicConfig(
    level=logging.DEBUG,
    format="%(asctime)s %(levelname)s %(message)s",
    handlers=[logging.FileHandler(LOG_FILE), logging.StreamHandler()],
)
logger = logging.getLogger(__name__)

# ── Year-detection regexes ──────────────────────────────────────────────────
_AID_YEAR_WORDS  = re.compile(r"Aid[\s]?Y(?:ea)?r|Year", re.IGNORECASE)
_TERM_WORD       = re.compile(r"Term", re.IGNORECASE)
_TERM_NUM        = re.compile(r"1[0-9][0-9][468]")
_DATE_RE         = re.compile(
    r"(0*[1-9]|1[012])[-/.](0*[1-9]|[12][0-9]|3[01])[-/.](2\d{3}|\d{2})"
    r"|(0*[1-9]|[12][0-9]|3[01])[-/.](0*[1-9]|1[012])[-/.](2\d{3}|\d{2})"
)
_INSTANCE_RE     = re.compile(r"[_][0-9]{2}[_-][0-9]+\.|[_-][0-9]+\.")
_YEAR_IN_CSV     = re.compile(r"[-_][0-9]{4}\.")
_YEAR_IN_NAME    = re.compile(r"_\d\d-")


class FileProcessor:
    """Processes one batch of files for a given aid year."""

    def __init__(self, year: str, is_test: bool) -> None:
        self.year    = year
        self.is_test = is_test

        self.date   = time.strftime("%x").replace("/", "-")
        self.month  = self.date[:2] + "-20" + self.date[-2:]
        self.aid_year_range = f"{int(year) - 1}-{year}"

        today = datetime.date.today()
        delta = 3 if today.weekday() == 0 else 1
        self.disb_date = (today - datetime.timedelta(days=delta)).strftime("%m-%d-%y")

        self.folder_path: Path = Path()
        self.direct_loan_flag  = False
        self.alt_loan_flag     = False
        self.unknown_list: list[str] = []
        self._query_dict: dict[str, str] = {}
        self._excel_cache: dict[str, tuple[bool, str]] = {}

        self._load_query_dict()
        logger.info(f"Initialized — year={year} test={is_test} disb_date={self.disb_date}")

    # ── Path helpers ───────────────────────────────────────────────────────

    @property
    def uosfa_dir(self) -> Path:
        return config.TEST_UOSFA_DIR if self.is_test else config.UOSFA_DIR

    def get_archive(self, key: str) -> Path:
        """Resolve an archive key to an absolute path for the current run."""
        y   = self.year
        ay  = self.aid_year_range
        mo  = self.month
        sys = config.TEST_SYSTEMS_DIR if self.is_test else config.SYSTEMS_DIR
        act = config.TEST_ACCT_DIR    if self.is_test else config.ACCT_DIR
        df  = config.TEST_DISB_FAIL   if self.is_test else config.DISB_FAIL_DIR

        # Monthly award-summary needs last month's name
        now        = datetime.datetime.now()
        last_month = now.month - 1 if now.month > 1 else 12
        last_year  = now.year - 1  if now.month == 1 else now.year
        award_sub  = (f"Award Summary {y}/"
                      f"Award Summary {calendar.month_name[last_month]} {last_year}")

        table: dict[str, Path] = {
            "daily":          sys / "QUERIES/Daily"              / ay / mo,
            "budget":         sys / "QUERIES/Budgets"            / ay / mo,
            "budget_wrong":   sys / "QUERIES/Budgets"            / ay / mo / "Wrong Budget Queries",
            "weekly":         sys / "QUERIES/Monday Weekly"      / ay / mo,
            "packaging":      sys / "QUERIES/Packaging"          / ay / mo,
            "monthly":        sys / "QUERIES/Monthly"            / mo,
            "monthly_dl":     sys / "Direct Loans/Monthly",
            "monthly_acct":   act / "Chartfields",
            "monthly_award":  act / "Award Summary" / award_sub,
            "sap":            sys / "QUERIES/SAP"                / y,
            "pell":           sys / "QUERIES/Pell Repackaging"   / ay,
            "pell_disb":      sys / "Pell Reports"               / ay / mo,
            "dl":             sys / "Direct Loans/DL Pre-Outbound",
            "dl_heal":        sys / "Direct Loans/DL HEAL Flag",
            "dl_response":    sys / "Direct Loans/DL Response Files",
            "dl_monthly":     sys / "Direct Loans/Monthly",
            "alt_loan":       sys / "QUERIES/ALT Loans",
            # Scholar paths differ between daily and weekly in the original
            "scholar_daily":  sys / f"{ay} Scholar/Queries",
            "scholar_weekly": (sys / "Scholarships" / f"{ay} Scholar/Queries"),
            "scholar_save":   sys / "Queries/Save" / y if not self.is_test
                              else config.TEST_SYSTEMS_DIR / "Save" / y,
            "scholar_errors": Path("C:/Systems/External Awards/Errors") if self.is_test
                              else sys / "External Awards/Errors",
            "ea":             sys / "External Awards/External Award Queries",
            "atb":            sys / "QUERIES/ATB",
            "3c":             sys / "QUERIES/3C Queries",
            "tsm":            Path("C:/QUERIES/TSM/NSLDS TSM Request") if self.is_test
                              else sys / "QUERIES/TSM/NSLDS TSM Request",
            "term":           sys / "QUERIES/Term"                / ay,
            "ldr":            sys / "QUERIES/LDR"                 / ay,
            "royall":         sys / "Royall",
            "save":           sys / "QUERIES/SAVE"                / ay,
            "disb":           sys / "QUERIES/Disbursement"        / ay / mo,
            "disb_failure":   df  / f"Disb Failure {ay}",
            "refund":         sys / "QUERIES/Refund Credit Holds" / mo,
            "pre_disb":       sys / "QUERIES/Disbursement/Pre-Disbursement Queries",
        }
        path = table.get(key)
        if path is None:
            raise KeyError(f"Unknown archive key: {key!r}")
        path.mkdir(parents=True, exist_ok=True)
        return path

    # ── Renaming helpers ───────────────────────────────────────────────────

    def _base_filename(self, name: str) -> str:
        """Strip date/year/instance markers, return core name + extension."""
        if name.endswith(".csv"):
            return name
        dot   = name.find(".")
        inst  = _INSTANCE_RE.search(name)
        inst_i = inst.start() if inst else -1
        has_y  = bool(_YEAR_IN_NAME.search(name))
        if has_y:
            return (name[:inst_i] if inst_i > -1 else name[: dot - 3]) + name[dot:]
        return (name[:inst_i] if inst_i > -1 else name[:dot]) + name[dot:]

    def new_name(self, name: str, year: str) -> str:
        base = self._base_filename(name)
        dot  = base.find(".")
        return f"{self.date} {base[:dot]} {year[2:]}{base[dot:]}"

    def new_name_disb(self, name: str, year: str) -> str:
        base = self._base_filename(name)
        dot  = base.find(".")
        return f"{self.disb_date} {base[:dot]} {year[2:]}{base[dot:]}"

    def _unique_path(self, folder: Path, filename: str) -> Path:
        """Return a path that does not already exist, adding (2), (3) … as needed."""
        p = folder / filename
        if not p.exists():
            return p
        stem, suffix = Path(filename).stem, Path(filename).suffix
        i = 2
        while True:
            candidate = folder / f"{stem} ({i}){suffix}"
            if not candidate.exists():
                return candidate
            i += 1

    # ── Aid-year detection ─────────────────────────────────────────────────

    def _has_aid_year_in_name(self, name: str) -> tuple[bool, str]:
        if ".csv" in name:
            m = _YEAR_IN_CSV.search(name)
            if m:
                return True, "20" + m.group()[-3:-1]
        m = _YEAR_IN_NAME.search(name)
        if m:
            return True, "20" + m.group()[1:-1]
        return False, "0"

    def _is_aid_year_word(self, value: object) -> bool:
        return bool(_AID_YEAR_WORDS.search(str(value)))

    def _is_aid_year_num(self, value: object) -> bool:
        s = str(value)
        m = re.search(r"[0-9]{2,4}[\s]*$", s)
        if not m:
            return False
        num = m.group(0).strip()
        n   = int(num)
        curr = int(self.year)
        if len(num) == 4:
            return curr - 5 < n < curr + 5
        if len(num) == 2:
            return (curr - 5) < (n + 2000) < (curr + 5 + 2000)
        return False

    def _parse_term_num(self, value: object) -> int:
        m = _TERM_NUM.search(str(value))
        if m:
            n = int(m.group()[1:3]) + 2000
            if int(self.year) - 5 < n < int(self.year) + 5:
                return n
        return -1

    def _search_excel_file(self, filename: Path) -> tuple[bool, str]:
        fullpath = self.folder_path / filename
        key = str(fullpath)
        if key in self._excel_cache:
            return self._excel_cache[key]

        try:
            with warnings.catch_warnings(record=True):
                warnings.simplefilter("always")
                with open(fullpath, "rb") as f:
                    buf = io.BytesIO(f.read())
                wb = openpyxl.load_workbook(buf, read_only=True, data_only=True)
            sheet = wb.active
            # Single streaming pass — never call .cell() on a read-only sheet
            rows = list(sheet.iter_rows(min_row=1, max_row=50, values_only=True))
            wb.close()
        except Exception as exc:
            logger.warning(f"Cannot read Excel {filename}: {exc}")
            result: tuple[bool, str] = (False, "0")
            self._excel_cache[key] = result
            return result

        has, year = False, "0"
        aid_cols: list[int] = []
        aid_rows: list[int] = []

        # Phase 1: scan first 5 rows for a header or inline year
        for ri, row in enumerate(rows[:5]):
            for ci, val in enumerate(row):
                if val is None:
                    break  # stop at first None in this row (matches original behaviour)
                sv = str(val)
                if self._is_aid_year_word(sv):
                    aid_cols.append(ci)
                    aid_rows.append(ri)
                    if self._is_aid_year_num(sv):
                        result = (True, "20" + sv[-2:])
                        self._excel_cache[key] = result
                        return result
                    if _DATE_RE.search(sv):
                        result = (True, "20" + sv[-2:])
                        self._excel_cache[key] = result
                        return result
                elif _TERM_WORD.search(sv):
                    ty = self._parse_term_num(sv)
                    if ty != -1:
                        logger.info(f"Year from Term header: {filename}")
                        result = (True, str(ty))
                        self._excel_cache[key] = result
                        return result

        # Phase 2: scan down each aid-year column for a max year value
        max_yr = 0
        for i in range(len(aid_cols)):
            col = aid_cols[i]
            for row in rows[aid_rows[i]:]:
                if row[0] is None:
                    break
                val = row[col] if col < len(row) else None
                if val is None:
                    continue
                sv = str(val)
                if self._is_aid_year_num(sv):
                    n = int(sv[-2:])
                    if n > max_yr:
                        max_yr, has, year = n, True, "20" + sv[-2:]
                elif _DATE_RE.search(sv):
                    n = int(sv[-2:])
                    if n > max_yr:
                        max_yr, has, year = n, True, "20" + sv[-2:]

        result = (has, year)
        self._excel_cache[key] = result
        return result

    def find_aid_year(self, filename: Path) -> str:
        """Return the aid year for this file (from its name or contents)."""
        s = str(filename)
        has, year = self._has_aid_year_in_name(s)
        if has:
            return year

        ext = s.lower()
        if ext.endswith(("xlsx", "xlsm", "xltx", "xltm")):
            found, year = self._search_excel_file(filename)
            return year if found else self.year

        return self.year

    # ── File processing ────────────────────────────────────────────────────

    def _do_query(self, name: str, renamed: str, archive: Path, uosfa_folder: str) -> None:
        """Copy file to archive; move to UOSFA folder (unless folder is None)."""
        src = self.folder_path / name
        try:
            if uosfa_folder == "None":
                dest = self._unique_path(archive, renamed)
                shutil.move(str(src), dest)
                logger.info(f"Archived (no UOSFA): {name} -> {dest}")
            else:
                uosfa_dest_dir = self.uosfa_dir / uosfa_folder
                uosfa_dest_dir.mkdir(parents=True, exist_ok=True)
                arc_dest   = self._unique_path(archive, renamed)
                uosfa_dest = self._unique_path(uosfa_dest_dir, renamed)
                shutil.copy(str(src), arc_dest)
                shutil.move(str(src), uosfa_dest)
                logger.info(f"Processed: {name} -> {uosfa_dest}")
        except Exception as exc:
            logger.error(f"Error processing {name}: {exc}")

    def do_query_unknown(self, name: str, renamed: str, folder: str, learn: bool) -> None:
        """Process a user-categorised unknown file; optionally save the mapping."""
        if learn and folder != "Unknown Reports":
            self._load_query_dict()
            self._query_dict[self._base_filename(name)] = folder
            self._save_query_dict()
            logger.info(f"Learned: {self._base_filename(name)} -> {folder}")

        src  = self.folder_path / name
        dest_dir = self.uosfa_dir / folder
        dest_dir.mkdir(parents=True, exist_ok=True)

        try:
            dest = self._unique_path(dest_dir, renamed)
            shutil.move(str(src), dest)
            logger.info(f"Unknown processed: {name} -> {dest}")
        except Exception as exc:
            logger.error(f"Error processing unknown {name}: {exc}")

    def process_file(self, filename: str, year: str) -> None:
        """Route one file: match against ROUTES, copy/move it."""
        stem = filename        # match against full filename (with extension)

        renamed      = self.new_name(filename, year)
        renamed_disb = self.new_name_disb(filename, year)

        # Delete files that should be removed entirely
        for pat in config.REMOVE_PATTERNS:
            if re.search(pat, filename):
                try:
                    os.remove(self.folder_path / filename)
                    logger.info(f"Deleted: {filename}")
                except Exception as exc:
                    logger.error(f"Delete failed {filename}: {exc}")
                return

        # Walk the routing table
        for rule in ROUTES:
            if rule.match(stem, year):
                if rule.rename_fn is not None:
                    renamed = rule.rename_fn(filename, year, self.date)
                elif rule.use_disb_date:
                    renamed = renamed_disb

                archive = self.get_archive(rule.archive)

                if rule.flag == "direct_loan":
                    self.direct_loan_flag = True
                elif rule.flag == "alt_loan":
                    self.alt_loan_flag = True

                self._do_query(filename, renamed, archive, rule.folder)
                return

        # Check the learned dictionary
        base = self._base_filename(filename)
        if base in self._query_dict:
            folder = self._query_dict[base]
            logger.info(f"Learned route for {filename}: {folder}")
            self.do_query_unknown(filename, renamed, folder, learn=False)
            return

        # Unknown — let the UI ask the user
        self.unknown_list.append(filename)
        logger.info(f"Unknown: {filename}")

    # ── Origination helpers ────────────────────────────────────────────────

    def _copy_orig(self, src_base: Path, dest_dir: Path, label: str) -> bool:
        dest_dir.mkdir(parents=True, exist_ok=True)
        copied = False
        for ext in (".doc", ".docx"):
            src  = src_base.parent / (src_base.name + ext)
            dest = dest_dir / (src_base.name + ext)
            if src.exists() and not dest.exists():
                try:
                    shutil.copy(src, dest)
                    logger.info(f"Copied {label}: {src} -> {dest}")
                    copied = True
                    break
                except Exception as exc:
                    logger.warning(f"Failed to copy {label} {src}: {exc}")
        return copied

    def move_direct_orig(self, filepath: Optional[Path] = None) -> bool:
        src_dir  = config.TEST_DL_ORIG_DIR   if self.is_test else config.DL_ORIG_DIR
        dest_dir = (config.TEST_UOSFA_DIR / "Direct Loan Reports" if self.is_test
                    else config.UOSFA_DIR / "Direct Loan Reports")
        base_name = f"{self.date} DL ORIG {self.year}"
        src_base  = (src_dir / base_name) if filepath is None else Path(filepath)
        ok = self._copy_orig(src_base, dest_dir, "DL ORIG")
        # Also attempt (2) variant
        self._copy_orig(src_base.parent / (src_base.name + " (2)"),
                        dest_dir, "DL ORIG (2)")
        return ok

    def move_alt_orig(self, filepath: Optional[Path] = None) -> bool:
        src_dir  = config.TEST_ALT_ORIG_DIR  if self.is_test else config.ALT_ORIG_DIR
        dest_dir = (config.TEST_UOSFA_DIR / "Alternative Loan Reports" if self.is_test
                    else config.UOSFA_DIR / "Alternative Loan Reports")
        base_name = f"{self.date} ALT Loan ORIG {self.year}"
        src_base  = (src_dir / base_name) if filepath is None else Path(filepath)
        return self._copy_orig(src_base, dest_dir, "ALT ORIG")

    # ── Dictionary persistence ─────────────────────────────────────────────

    def _load_query_dict(self) -> None:
        try:
            with open(config.DICT_PATH, newline="") as f:
                for row in csv.reader(f):
                    if len(row) >= 2:
                        self._query_dict[row[0]] = row[1]
            logger.debug(f"Dict loaded: {len(self._query_dict)} entries")
        except FileNotFoundError:
            pass
        except Exception as exc:
            logger.error(f"Dict load error: {exc}")

    def _save_query_dict(self) -> None:
        try:
            with open(config.DICT_PATH, "w", newline="\n") as f:
                csv.writer(f).writerows(self._query_dict.items())
            logger.debug(f"Dict saved: {len(self._query_dict)} entries")
        except Exception as exc:
            logger.error(f"Dict save error: {exc}")

    # ── Main entry points ──────────────────────────────────────────────────

    def set_source_folder(self, directory: str) -> None:
        self.folder_path = Path(directory)

    def process_all(self) -> None:
        """Process every file in folder_path."""
        if not self.folder_path or not self.folder_path.is_dir():
            logger.error("No valid source folder set")
            return

        files = [f for f in os.listdir(self.folder_path)
                 if os.path.isfile(self.folder_path / f)]
        logger.info(f"Source: {self.folder_path}  ({len(files)} files)")

        for i, fname in enumerate(files, 1):
            if any(skip in fname for skip in config.SKIP_CONTAINS):
                logger.debug(f"Skipped: {fname}")
                continue
            logger.info(f"[{i}/{len(files)}] {fname}")
            year = self.find_aid_year(Path(fname))
            self.process_file(fname, year)

        logger.info(
            f"Done — direct_loan={self.direct_loan_flag} "
            f"alt_loan={self.alt_loan_flag} "
            f"unknown={len(self.unknown_list)}"
        )

    def run(self) -> tuple[bool, bool, list[str]]:
        """
        Ask the user to pick a source folder, process all files in it,
        and return (direct_loan_flag, alt_loan_flag, unknown_list).
        This method does NOT open a tkinter dialog; the caller must call
        set_source_folder() first.
        """
        self.process_all()
        return self.direct_loan_flag, self.alt_loan_flag, self.unknown_list
