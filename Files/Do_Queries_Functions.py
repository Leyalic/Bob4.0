# Created by Joshua Hardy, mmason, and Iman
from genericpath import isfile
import os
from pickle import FALSE
import re
#import xlrd
import openpyxl
import datetime
import time
import shutil
import tkinter
from tkinter import filedialog
from pathlib import Path
import csv
import warnings
import io
import logging

# Query imports
import sys
sys.path.insert(1, '../Files/')
from Files import After_Repack_Queries
from Files import Alt_Loan_Queries
from Files import Atb_Fbill_3C_Queries
from Files import Budget_Queries
from Files import Daily_Queries
from Files import Day_AfterLDR
from Files import Direct_Loan
from Files import Disbursement_Queries
from Files import EndOfTerm_Queries
from Files import Mid_Repack_Queries
from Files import Monday_WeeklyQueries
from Files import Monthly_Queries
from Files import Packaging_Queries
from Files import PrePackaging_Queries
from Files import Scholarships_Queries
from Files import Second_LDR
from Files import Tsm_Queries

# Configure logging with both console and file output
_log_dir = Path("./logs")
_log_dir.mkdir(exist_ok=True)
_log_file = _log_dir / f"bob_queries_{time.strftime('%Y%m%d_%H%M%S')}.log"

logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(_log_file),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)
logger.info(f"Logging initialized - Log file: {_log_file}")


global test

skip_files = [".zip", ".lnk", " DL ORIG ", " ALT Loan ORIG "]
remove_files = ["FASTDVER", "FINAID_Checklist", "ussfa09", "USSFA090 Reset", "O-A", r"uosfa[0-9]+[a-zA-Z]*\.csv"]

# The date becomes the current date and is then placed in MM-DD-YY format
date = time.strftime("%x").replace("/", "-")
now = datetime.datetime.now()
last_month = now.month - 1 if now.month > 1 else 12
last_months_year = now.year - 1 if now.month == 12 else now.year
month_folder = date[:2] + "-20" + date[-2:]

# Var definitions
odd_aid_years = []
even_aid_years = []
unknown_list = []
current_aid_year = ""
folder_path = Path()
disbursement_date = datetime.datetime.min
alt_loan_flag = False
direct_loan_flag = False
excel_parse_cache = {}  # Cache for Excel file parsing to avoid re-opening files

# Regular Expressions
aid_year_regex = ["Aid[\s]?Y(ea)?r", "Year"]
aid_num_regex = ["[0-9]{2,4}[\s]*$"]
term_regex = ["Term"]
term_num_regex = ["1[0-9][0-9][468]"]
date_regex = ["(0*[1-9]|1[012])[-/.](0*[1-9]|[12][0-9]|3[01])[-/.](2\d{3}|\d{2})","(0*[1-9]|[12][0-9]|3[01])[-/.](0*[1-9]|1[012])[-/.](2\d{3}|\d{2})"]
instance_regex = ["[_][0-9]{2}[_-][0-9]+\.", "[_-][0-9]+\."]

# Directories
#test_UOSFA_directory = Path("C:/Users/iessaghir/Documents/DoQueries/Destination Folders")
#test_UOSFA_directory = Path("C:/Users/JHARDY/Documents/DoQueries/Destination Folders")
test_UOSFA_directory = Path("C:/UOSFA Reports/Testing/Destination Folders")

UOSFA_directory = Path("O:/UOSFA Reports")

# Query Routing Dictionary - maps query functions and metadata
# Format: (name, function, [arg_keys], sets_flag)
# arg_keys reference values from the local scope in move_files()
QUERY_MODULES = [
    ('Daily Queries', Daily_Queries.do_dailies, ['test', 'date', 'year', 'filename', 'renamed'], None),
    ('Monday Weekly', Monday_WeeklyQueries.do_monday_weeklies, ['test', 'date', 'year', 'filename', 'renamed'], None),
    ('Budget Queries', Budget_Queries.do_budget_queries, ['test', 'date', 'year', 'filename', 'renamed'], None),
    ('Packaging Queries', Packaging_Queries.do_packaging_queries, ['test', 'date', 'year', 'filename', 'renamed'], None),
    ('Monthly Queries', Monthly_Queries.do_monthlies, ['test', 'date', 'year', 'filename', 'renamed'], None),
    ('Disbursement Queries', Disbursement_Queries.do_disb_queries, ['test', 'date', 'year', 'filename', 'renamed', 'renamed_disb'], None),
    ('2nd LDR', Second_LDR.do_2nd_ldr, ['test', 'year', 'filename', 'renamed'], None),
    ('End of Term', EndOfTerm_Queries.do_end_of_term_queries, ['test', 'date', 'year', 'filename', 'renamed'], None),
    ('Day After LDR', Day_AfterLDR.do_day_after_ldr, ['test', 'year', 'filename', 'renamed'], None),
    ('Direct Loan Pre-Outbound', Direct_Loan.dl_pre_outbound, ['test', 'date', 'year', 'filename', 'renamed'], 'direct_loan_flag'),
    ('Alt Loan Pre-Outbound', Alt_Loan_Queries.al_pre_outbound, ['test', 'filename', 'renamed'], 'alt_loan_flag'),
    ('Pre-Repackaging', PrePackaging_Queries.do_pre_repackaging, ['test', 'year', 'filename', 'renamed'], None),
    ('Mid-Repackaging', Mid_Repack_Queries.do_mid_repack_queries, ['test', 'year', 'filename', 'renamed'], None),
    ('After Repackaging', After_Repack_Queries.do_after_repackaging, ['test', 'year', 'filename', 'renamed'], None),
    ('Daily Scholarships', Scholarships_Queries.do_daily_scholarships, ['test', 'year', 'filename', 'renamed'], None),
    ('Weekly Scholarships', Scholarships_Queries.do_weekly_scholarships, ['test', 'year', 'filename', 'renamed'], None),
    ('Budget Testing', Budget_Queries.do_budget_test_queries, ['test', 'date', 'year', 'filename', 'renamed'], None),
    ('ATB and 3C', Atb_Fbill_3C_Queries.do_atb_fb_3c_queries, ['test', 'filename', 'renamed'], None),
    ('Transfer Student Monitoring', Tsm_Queries.do_tsm_queries, ['test', 'filename', 'renamed'], None),
]

# Origination Filepaths
test_dir_orig_folder = Path('C:/Testing Bob/Direct Loans/Origination')
test_alt_orig_folder = Path('C:/Testing Bob/ALT Loans/')
dir_orig_folder = Path('O:/Systems/Direct Loans/Origination')
alt_orig_folder = Path('O:/Systems/QUERIES/ALT Loans')

# Automatically Updated Queries
query_dict = {} # Data format =  {"Query Name" : "UOSFA Folder"}
dict_path = "./Files/Query_Dictionary.csv"

# Odd year
def is_odd_year(year):
    year_int = int(year[-1])
    return year_int % 2 == 1

def is_aid_year_word(value):
    return any (re.search(regex_str, str(value), re.IGNORECASE) for regex_str in aid_year_regex)

def is_term_word(value):
    return any (re.search(regex_str, str(value), re.IGNORECASE) for regex_str in term_regex)

def is_aid_year_num(value):
    # Find 2 to 4 digit number
    result = re.search(aid_num_regex[0], str(value))
    
    # Check if number is within aid year range
    if result:
        match = result.group(0)
        match_num = int(match)
        if len(match) == 4:
            curr_num = int(current_aid_year)
            return (match_num > curr_num - 5) and (match_num < curr_num + 5)
        if len(match) == 3:
            return False
        if len(match) == 2:
            curr_num = int(str(current_aid_year)[-2:]) + 2000
            match_num += 2000
            return (match_num > curr_num - 5) and (match_num < curr_num + 5)

    return result

def parse_term_num(value):
    # Find term formatted number
    result = re.search(term_num_regex[0], str(value))
    
    # Check if number is within aid year range
    if result:
        match = result.group(0)
        match_num = int(match[1:3])
        curr_num = int(current_aid_year)
        match_num += 2000
        if (match_num > curr_num - 5) and (match_num < curr_num + 5):
            return match_num

    return -1


def is_date(value):
    return any (re.search(regex_str, str(value)) for regex_str in date_regex)


# Search for and return aid year in filename
def has_aid_year(filename):
    filestring = str(filename)
    has = False
    year = "0"

    if ".csv" in filestring:
        found_year = re.search(r"[-_][0-9]{4}\.", filestring)
        if found_year:
            has = True
            year = "20" + found_year.group()[-3:-1]
            return (has, year)

    found_year = re.search(r"_\d\d-", filestring)
    if found_year:
        has = True
        year = "20" + found_year.group()[1:-1]
    return (has, year)


# Finds the aid year of the file either in the file's name or contents
def find_aid_year(filename):
    global current_aid_year
    global odd_aid_years
    global even_aid_years

    aid_year = current_aid_year # Default is current year
    filestring = str(filename)  
    file_year = has_aid_year(filestring) # file_year: (bool has_year, str year)

    # Check in filename
    if file_year[0]:
        aid_year = file_year[1]

    # Check in .xls cells - Retired as of Aug 2, 2024, xlrd no longer supported
    #elif filestring.lower().endswith("xls"):
       # file_year = search_xls_file(filename)
       # if file_year[0]:
      #      aid_year = file_year[1]
      #  else: 
      #      aid_year = current_aid_year
    
    # Check in other spreadsheet cells
    elif filestring.lower().endswith(('xlsx', 'xlsm', 'xltx', 'xltm')):
        file_year = search_excel_file(filename) 
        if file_year[0]:
            aid_year = file_year[1]
        else: 
            aid_year = current_aid_year

    # Record file name in appropriate list
    if is_odd_year(aid_year):
        odd_aid_years.append(filestring)
    else:
        even_aid_years.append(filestring)

    return aid_year


# Search for and return maximum aid year in xls file
# Changes as of Aug 2, 2024 :  Commenting the xlrd related search query as the format is no longer supported
#def search_xls_file(filename):
 #   global folder_path
  #  has = False
   # year = "0"

    #fullpath = folder_path / filename
    #workbook = xlrd.open_workbook(fullpath, logfile=open(os.devnull, 'w'))
    #sheet = workbook.sheet_by_index(0)
    #aid_cols = []
    #aid_rows = []
    #max_year = 0
    
    # Locate cells containing the word 'Aid Year'
    #for row in range(sheet.nrows):
     #   if row > 5:
            #print(" Quit Early: " + str(filename))
      #      break
       # for col in range(sheet.ncols):
        #    value = sheet.cell_value(row, col)
         #   if is_aid_year_word(value):
          #      aid_cols.append(col)
           #     aid_rows.append(row)
            #    if is_aid_year_num(value):
             #       has = True
              #      year = "20" +  str(value)[-2:] 
               #     workbook.release_resources()
                #    return (has, year)
                #elif is_date(value):
                 #   has = True
                  #  year = "20" +  str(value)[-2:] # Assumes date format w/ year at end
                   # workbook.release_resources()
                    #return (has, year)

            # Sometimes there's a "Term" value instead of an "Aid Year"
            #elif is_term_word(value):
             #   term_year = parse_term_num(value)
              #  if term_year != -1:
               #     has = True
                #    year = str(term_year)
                 #   workbook.release_resources()
                  #  print("Used Term instead of Aid Year: " + str(filename))
                   # return (has, year)

    
    # Find maximum aid year in 'Aid Year' column
   # for i in range(len(aid_cols)):
    #    aid_col = aid_cols[i]
     #   aid_row = aid_rows[i]
      #  if aid_col > -1:
       #     curr_row = aid_row
        #    while (curr_row < sheet.nrows):
         #       value = str(sheet.cell_value(curr_row, aid_col))
          #      if is_aid_year_num(value):
           #         value_int = int(value[-2:])
            #        if value_int > max_year:
             #           max_year = value_int
              #          has = True
               #         year = "20" +  str(value)[-2:]
                #elif is_date(value):
                 #   value_int = int(value[-2:])
                  #  if value_int > max_year:
                   #     max_year = value_int
                    #    has = True
                     #   year = "20" +  str(value)[-2:] # Assumes date format w/ year at end
                #curr_row = curr_row + 1

    #workbook.release_resources()
    # return (has, year)


# Search for and return aid year in excel file
def search_excel_file(filename):
    global folder_path

    fullpath = folder_path / filename
    cache_key = str(fullpath)
    if cache_key in excel_parse_cache:
        return excel_parse_cache[cache_key]

    has = False
    year = "0"

    try:
        with warnings.catch_warnings(record=True):
            warnings.simplefilter("always")
            with open(fullpath, "rb") as f:
                in_mem_file = io.BytesIO(f.read())
            workbook = openpyxl.load_workbook(in_mem_file, read_only=True, data_only=True)

        sheet = workbook.active

        # Read first 50 rows in a single streaming pass — far faster than repeated .cell() calls
        all_rows = list(sheet.iter_rows(min_row=1, max_row=50, values_only=True))
        workbook.close()

    except Exception as e:
        logger.warning(f"Error reading Excel file {filename}: {e}")
        result = (False, "0")
        excel_parse_cache[cache_key] = result
        return result

    aid_cols = []
    aid_rows = []
    max_year = 0

    # Phase 1: scan first 5 rows for an Aid Year / Term header or inline year
    for row_idx, row in enumerate(all_rows[:5]):
        for col_idx, value in enumerate(row):
            if value is None:
                break  # original behaviour: stop at first None in a row
            strval = str(value)
            if is_aid_year_word(strval):
                aid_cols.append(col_idx)
                aid_rows.append(row_idx)
                if is_aid_year_num(strval):
                    has = True
                    year = "20" + strval[-2:]
                    result = (has, year)
                    excel_parse_cache[cache_key] = result
                    return result
                elif is_date(value):
                    has = True
                    year = "20" + strval[-2:]
                    result = (has, year)
                    excel_parse_cache[cache_key] = result
                    return result
            elif is_term_word(strval):
                term_year = parse_term_num(strval)
                if term_year != -1:
                    has = True
                    year = str(term_year)
                    logger.info(f"Used Term instead of Aid Year: {filename}")
                    result = (has, year)
                    excel_parse_cache[cache_key] = result
                    return result

    # Phase 2: scan downward in each column that had an 'Aid Year' header
    for i in range(len(aid_cols)):
        aid_col = aid_cols[i]
        for row in all_rows[aid_rows[i]:]:
            if row[0] is None:  # col 1 None signals end of data rows
                break
            value = row[aid_col] if aid_col < len(row) else None
            if value is None:
                continue
            strval = str(value)
            if is_aid_year_num(strval):
                value_int = int(strval[-2:])
                if value_int > max_year:
                    max_year = value_int
                    has = True
                    year = "20" + strval[-2:]
            elif is_date(strval):
                value_int = int(strval[-2:])
                if value_int > max_year:
                    max_year = value_int
                    has = True
                    year = "20" + strval[-2:]

    result = (has, year)
    excel_parse_cache[cache_key] = result
    return result


# Prints list of sorted files to terminal (debugging)
def output_sorted_files():
    """Log summary of how files were categorized."""
    logger.info(f"Processing Summary - Odd Years: {len(odd_aid_years)}, Even Years: {len(even_aid_years)}, Unknown: {len(unknown_list)}")
    
    if odd_aid_years:
        logger.debug(f"Odd Years: {odd_aid_years}")
    if even_aid_years:
        logger.debug(f"Even Years: {even_aid_years}")
    if unknown_list:
        logger.info(f"Unknown files ({len(unknown_list)}): {unknown_list}")


# Renames file to ensure no duplicates at desired location
def rename_no_duplicates(folder_path, renamed):
    """Ensure unique filename in destination folder using Path objects."""
    folder_path = Path(folder_path)
    filepath = folder_path / renamed
    
    if not filepath.exists():
        return str(filepath)
    
    # File exists, need to add counter
    stem = filepath.stem  # filename without extension
    suffix = filepath.suffix  # file extension
    parent = filepath.parent
    counter = 2
    
    while True:
        new_filename = f"{stem} ({counter}){suffix}"
        new_filepath = parent / new_filename
        if not new_filepath.exists():
            return str(new_filepath)
        counter += 1

# Renames, copies, and moves file to desired destination
def do_query(name, renamed, legacy_archive, UOSFA_folder, year):
    """Process file: copy to archive, move to UOSFA folder."""
    current_filepath = Path(folder_path) / name
    
    if test:
        UOSFA_destination = test_UOSFA_directory / UOSFA_folder
    else:       
        UOSFA_destination = UOSFA_directory / UOSFA_folder
    
    # Ensure destination folder exists
    UOSFA_destination.mkdir(parents=True, exist_ok=True)
    
    try:
        if UOSFA_folder == "None":
            # Just move to archive, don't copy to UOSFA
            legacy_filepath = rename_no_duplicates(legacy_archive, renamed)
            shutil.move(str(current_filepath), legacy_filepath)
            logger.info(f"Moved {name} to {legacy_filepath}")
        else:
            # Copy to archive, then move to UOSFA
            legacy_filepath = rename_no_duplicates(legacy_archive, renamed)
            UOSFA_filepath = rename_no_duplicates(UOSFA_destination, renamed)
            shutil.copy(str(current_filepath), legacy_filepath)
            shutil.move(str(current_filepath), UOSFA_filepath)
            logger.info(f"Processed {name} -> {UOSFA_filepath}")
    except Exception as e:
        logger.error(f"Error processing {name}: {e}")
 

# Returns the corresponding archive folder associated with the given UOSFA folder
def get_unknown_archive(UOSFA_folder):
    """Map UOSFA folder to archive location using Path objects."""
    aid_year = f"{int(current_aid_year) - 1}-{current_aid_year}"
    month_folder = date[:2] + "-20" + date[-2:]
    
    archive_map = {
        "Alternative Loan Reports": Path('O:/Systems/QUERIES/ALT Loans/'),
        "Budget Reports": Path('O:/Systems/QUERIES/Budgets') / aid_year / month_folder,
        "Daily Reports": Path('O:/Systems/QUERIES/Daily') / aid_year / month_folder,
        "Direct Loan Reports": Path('O:/Systems/Direct Loans/DL Pre-Outbound'),
        "External Award Reports": Path('O:/Systems/External Awards/External Award Queries'),
        "Financial Aid Reports": Path('O:/Systems/QUERIES/'),
        "Monthly Reports": Path('O:/Systems/QUERIES/Monthly') / month_folder,
        "Packaging Reports": Path('O:/Systems/QUERIES/Packaging') / aid_year / month_folder,
        "Pell Reports": Path('O:/Systems/QUERIES/Pell Repackaging') / aid_year,
        "SAP Reports": Path('O:/Systems/QUERIES/SAP/') / current_aid_year,
        "Scholarship Reports": Path('O:/Systems/Scholarships') / f"{aid_year} Scholar/Queries",
        "Weekly Reports": Path('O:/Systems/QUERIES/Monday Weekly') / aid_year / month_folder,
    }
    
    # Return None for "Other Reports" and "Unknown Reports" (no archive)
    if UOSFA_folder in ("Other Reports", "Unknown Reports"):
        return None
    
    archive = archive_map.get(UOSFA_folder)
    if archive:
        archive.mkdir(parents=True, exist_ok=True)
        return str(archive)
    
    logger.warning(f"Unknown UOSFA folder: {UOSFA_folder}")
    return None


# Renames and moves file to manually selected folder
def do_query_unknown(name, renamed, destination, year, add_query):
    """Process unknown file: user selects destination folder."""
    global query_dict
    
    # Add query destination to query dictionary for future runs
    if add_query:      
        if destination != "Unknown Reports":
            load_dictionary()
            cleaned = clean_filename(name)
            query_dict[cleaned] = destination
            save_dictionary()
            logger.info(f"Learned: {cleaned} -> {destination}")
    
    # Copy to archive and move to UOSFA folder
    current_filepath = Path(folder_path) / name
    
    if test:
        UOSFA_destination = test_UOSFA_directory / destination
        archive = None
    else:       
        UOSFA_destination = UOSFA_directory / destination
        archive = get_unknown_archive(destination)
    
    UOSFA_destination.mkdir(parents=True, exist_ok=True)
    
    try:
        if archive is not None:
            archive_filepath = rename_no_duplicates(archive, renamed)
            shutil.copy(str(current_filepath), archive_filepath)
            logger.info(f"Copied to archive: {archive_filepath}")
        
        destination_filepath = rename_no_duplicates(UOSFA_destination, renamed)
        shutil.move(str(current_filepath), destination_filepath)
        logger.info(f"Unknown file processed: {name} -> {destination}")
    except Exception as e:
        logger.error(f"Error processing unknown file {name}: {e}")



# Extracts base filename (without aid year, date, or instance suffix)
def get_base_filename(filename):
    """Extract core filename by removing aid year and instance markers."""
    if filename.endswith('.csv'):
        return filename
    
    dot_index = filename.find(".")
    res = get_regex_result(filename, instance_regex)
    instance_index = -1
    if res is not None:
        instance_index = res.start()
    
    has_year, _ = has_aid_year(filename)
    
    if has_year:
        if instance_index > -1:
            return filename[:instance_index] + filename[dot_index:]
        else:
            return filename[:(dot_index - 3)] + filename[dot_index:]
    else:
        if instance_index > -1:
            return filename[:instance_index] + filename[dot_index:]
        else:
            return filename[:dot_index] + filename[dot_index:]

# Clean filename for dictionary lookup (remove date/year info)
def clean_filename(filename):
    """Alias for get_base_filename - used for query dictionary lookups."""
    return get_base_filename(filename)

def get_regex_result(word, regex_list):
    for i in range(len(regex_list)):
        res = re.search(regex_list[i], word)
        if res:
            return res
    return None

# Returns renamed filename with optional date prefix and aid year included
def _format_renamed_filename(name, year, prefix=""):
    """Core renaming logic: adds date/time prefix and year suffix."""
    base = get_base_filename(name)
    # base already has the extension, so extract it and rebuild
    dot_index = base.find(".")
    name_part = base[:dot_index]
    extension = base[dot_index:]
    
    if prefix:
        return f"{prefix} {name_part} {year[2:]}{extension}"
    else:
        return f"{name_part} {year[2:]}{extension}"

# Returns renamed filename with current date and aid year included        
def new_name(name, year):
    """Rename file with current date and aid year."""
    return _format_renamed_filename(name, year, date)

# Returns renamed filename with disbursement date and aid year included   
def new_name_disb(name, year):
    """Rename file with disbursement date and aid year."""
    return _format_renamed_filename(name, year, disbursement_date)


# Checks query files to determine the arguments to do_query
def move_files(filename, year):
    global current_aid_year
    global odd_aid_years
    global even_aid_years
    global unknown_list
    global disbursement_date
    global alt_loan_flag
    global direct_loan_flag

    date = time.strftime("%x").replace("/", "-")
    renamed = new_name(filename, year)
    renamed_disb = new_name_disb(filename, year)
    
    info = "Empty"  # Stores do_query parameters
    
    # Handle files that should be removed
    if any(re.search(regex_str, filename) for regex_str in remove_files):
        try:
            os.remove(folder_path / Path(filename))
            logger.info(f"Removed {filename}")
            if is_odd_year(year):
                odd_aid_years.remove(filename)
            else:
                even_aid_years.remove(filename)
        except Exception as e:
            logger.error(f"Failed to remove {filename}: {e}")
        return
    
    # Route through query modules using the routing dictionary
    local_vars = {
        'test': test,
        'date': date,
        'year': year,
        'filename': filename,
        'renamed': renamed,
        'renamed_disb': renamed_disb,
    }
    
    for query_name, query_func, arg_keys, flag_name in QUERY_MODULES:
        if info == "Empty":
            try:
                # Build arguments based on arg_keys
                args = [local_vars[key] for key in arg_keys]
                info = query_func(*args)
                
                if info != "Empty" and flag_name:
                    # Set flag if this query requires it
                    if flag_name == 'direct_loan_flag':
                        direct_loan_flag = True
                    elif flag_name == 'alt_loan_flag':
                        alt_loan_flag = True
                    logger.info(f"Matched {query_name}: {filename}")
            except Exception as e:
                logger.warning(f"Error in {query_name}: {e}")
                info = "Empty"  # Continue to next query
    
    # Check query dictionary for previously learned files
    if info == "Empty":
        cleaned = clean_filename(filename)
        if cleaned in query_dict:
            val = query_dict[cleaned]
            logger.info(f"Using learned folder for {filename}: {val}")
            do_query_unknown(filename, renamed, val, year, False)
            return
    
    # Handle unknown or processed files
    if info == "Empty":
        unknown_list.append(str(filename))
        logger.info(f"Unknown file: {filename}")
    else:
        do_query(info[0], info[1], info[2], info[3], year)


# Copy origination file for direct loans
# Helper function to copy origination files (handles .doc and .docx)
def _copy_orig_file_variants(source_base, dest_base, description):
    """Try to copy file with .doc and .docx variants."""
    source_base = Path(source_base)
    dest_base = Path(dest_base)
    dest_base.parent.mkdir(parents=True, exist_ok=True)
    
    variants = ['.doc', '.docx']
    copied = False
    
    for variant in variants:
        source = source_base.parent / (source_base.name + variant)
        dest = dest_base.parent / (dest_base.name + variant)
        
        if source.exists() and not dest.exists():
            try:
                shutil.copy(source, dest)
                logger.info(f"Copied {description}: {source} -> {dest}")
                copied = True
                break
            except Exception as e:
                logger.warning(f"Failed to copy {source}: {e}")
    
    return copied


# Copy origination file for direct loans
def move_direct_orig(filepath, dflt):
    """Copy Direct Loan origination files (.doc/.docx variants)."""
    if test:
        source_folder = test_dir_orig_folder
        dest_folder = test_UOSFA_directory / "Direct Loan Reports"
    else:
        source_folder = dir_orig_folder
        dest_folder = Path("O:/UOSFA Reports/Direct Loan Reports")
    
    base_name = f"{date} DL ORIG {current_aid_year}"
    
    try:
        if dflt:
            # Try default location first
            source_base = source_folder / base_name
        else:
            # Use user-provided path
            source_base = Path(filepath)
        
        dest_folder.mkdir(parents=True, exist_ok=True)
        
        # Copy main file and variant
        main_copied = _copy_orig_file_variants(source_base, dest_folder / base_name, "DL ORIG")
        
        # Try to copy (2) variant if exists
        source_variant = source_base.parent / f"{source_base.name} (2)"
        variant_copied = _copy_orig_file_variants(source_variant, dest_folder / f"{base_name} (2)", "DL ORIG (2)")
        
        return main_copied or variant_copied
    except Exception as e:
        logger.error(f"Error in move_direct_orig: {e}")
        return False


# Copy origination file for alt loans
def move_alt_orig(filepath, dflt):
    """Copy Alternative Loan origination files (.doc/.docx variants)."""
    if test:
        source_folder = test_alt_orig_folder
        dest_folder = test_UOSFA_directory / "Alternative Loan Reports"
    else:
        source_folder = alt_orig_folder
        dest_folder = Path("O:/UOSFA Reports/Alternative Loan Reports")
    
    base_name = f"{date} ALT Loan ORIG {current_aid_year}"
    
    try:
        if dflt:
            source_base = source_folder / base_name
        else:
            source_base = Path(filepath)
        
        dest_folder.mkdir(parents=True, exist_ok=True)
        main_copied = _copy_orig_file_variants(source_base, dest_folder / base_name, "ALT ORIG")
        
        return main_copied
    except Exception as e:
        logger.error(f"Error in move_alt_orig: {e}")
        return False

    
# Sort all files in directory based on aid year, and send to corresponding folder
def sort_files():
    global folder_path
    global test
    global unknown_list

    root = tkinter.Tk()    
    root.withdraw()
    directory = filedialog.askdirectory()
    root.destroy()
    
    if directory == "":
        logger.warning("No directory selected")
        return

    # Skip nested folders
    folder_path = directory
    files = [filepath for filepath in os.listdir(directory) if os.path.isfile(Path(directory) / Path(filepath))]
    
    logger.info(f"Processing directory: {directory}")
    logger.info(f"Found {len(files)} files to process")
    
    skipped_count = 0
    for i, filename in enumerate(files, 1):
        # Skip certain files
        if any(word in filename for word in skip_files):
            logger.debug(f"Skipping file (blacklisted): {filename}")
            skipped_count += 1
            continue

        # Send file to corresponding folder
        logger.info(f"Processing [{i}/{len(files)}]: {filename}")
        pFilename = Path(filename)
        aid_year = find_aid_year(pFilename)  
        move_files(filename, aid_year)
    
    logger.info(f"Skipped {skipped_count} files, processed {len(files) - skipped_count} files")


# Saves the Query Dictionary to a .csv file
def save_dictionary():
    try:
        with open(dict_path, "w", newline="\n") as data:
            w = csv.writer(data)
            for key, value in query_dict.items():
                w.writerow([key, value])
        logger.debug(f"Dictionary saved: {len(query_dict)} entries")
    except Exception as e:
        logger.error(f"Error saving dictionary: {e}")

# Loads the Query Dictionary from a .csv file
def load_dictionary():
    global query_dict
    try:
        with open(dict_path) as data:
            reader = csv.reader(data)
            for rows in reader:
                if rows:
                    key = rows[0]
                    value = rows[1]
                    query_dict[key] = value
        logger.debug(f"Dictionary loaded: {len(query_dict)} entries")
    except FileNotFoundError:
        logger.info("No existing query dictionary found")
    except Exception as e:
        logger.error(f"Error loading dictionary: {e}")

def test_save():
    testdict = {"query": "folder", "UFAA": "Budget Reports"}
    try:
        with open(dict_path, "w", newline="\n") as data:
            w = csv.writer(data)
            for key, value in testdict.items():
                w.writerow([key, value])
        logger.debug("Test dictionary saved")
    except Exception as e:
        logger.error(f"Error in test_save: {e}")

def test_load():
    testdict = {}
    try:
        with open(dict_path) as data:
            reader = csv.reader(data)
            for rows in reader:
                key = rows[0]
                value = rows[1]
                testdict[key] = value
        logger.debug(f"Test dictionary loaded: {len(testdict)} entries")
    except Exception as e:
        logger.error(f"Error in test_load: {e}")


# Reset global variables and initialize with user input
def initialize(year, is_test):
    global current_aid_year
    global STerm
    global disbursement_date
    global alt_loan_flag
    global direct_loan_flag
    global folder_path
    global test
    global odd_aid_years
    global even_aid_years
    global unknown_list

    odd_aid_years = []
    even_aid_years = []
    unknown_list = []

    test = is_test

    folder_path = Path()

    alt_loan_flag = False
    direct_loan_flag = False

    current_aid_year = year
    today = datetime.date.today()
    if today.weekday() == 0:
        disbursement_date = today - datetime.timedelta(days=3)
        disbursement_date = disbursement_date.strftime("%m-%d-%y")
    else:
        disbursement_date = today - datetime.timedelta(days=1)
        disbursement_date = disbursement_date.strftime("%m-%d-%y")
    
    logger.info(f"Initialized - Aid Year: {year}, Test Mode: {is_test}, Disbursement Date: {disbursement_date}")
    load_dictionary()

def run(year, is_test):
    logger.info(f"Starting file processing - Aid Year: {year}, Test Mode: {is_test}")
    
    initialize(year, is_test)
    sort_files()
    output_sorted_files()
    
    logger.info("File processing complete")
    return (direct_loan_flag, alt_loan_flag, unknown_list)
        



